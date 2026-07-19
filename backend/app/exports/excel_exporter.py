"""Excel (.xlsx) exporter using openpyxl."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database.models.generation_run import GenerationRun
from app.exports.base import Exporter, ExportResult

COLUMNS = [
    "Test Case ID", "Requirement ID", "Scenario", "Objective", "Priority",
    "Severity", "Preconditions", "Test Data", "Steps", "Expected Result",
    "Post Conditions", "Automation Candidate", "Automation Type", "Risk Level",
    "Test Type",
]


class ExcelExporter(Exporter):
    format_name = "excel"

    def export(self, run: GenerationRun) -> ExportResult:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Test Cases"

        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, header in enumerate(COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, tc in enumerate(run.test_cases, start=2):
            values = [
                tc.test_case_key, tc.requirement_traceability, tc.scenario, tc.objective,
                tc.priority.value, tc.severity.value, tc.preconditions,
                str(tc.test_data) if tc.test_data else "",
                "\n".join(f"{i}. {s}" for i, s in enumerate(tc.steps, start=1)),
                tc.expected_result, tc.post_conditions,
                "Yes" if tc.is_automation_candidate else "No", tc.automation_type,
                tc.risk_level.value, tc.test_type.value,
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")

        for col_idx in range(1, len(COLUMNS) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 24
        sheet.freeze_panes = "A2"

        buffer = io.BytesIO()
        workbook.save(buffer)
        return ExportResult(
            filename=f"testcases_{run.id}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=buffer.getvalue(),
        )
